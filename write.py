"""
Notice:
Some aspects of the following program are hardcoded.
It will most likely require re-editing depending on datasets in the future.
"""

from openpyxl import Workbook
from custom_dataclass import *

filename = "export/system_2022_v02.xlsx"
workbook = Workbook()
sheet = workbook.active


def print_rows() -> None:
    """
    Prints all rows in the sheet.
    :return: None
    """
    for row in sheet.iter_rows(values_only=True):
        print(row)


def prepare_sheet(days: int, worksites: list[Worksite]) -> None:
    """
    Prepares the sheet for writing. Adds headers and stuff...
    :param days: total number of columns
    :param worksites: total number of rows
    :return: None
    """
    for iteration in range(1, (days+1)):
        sheet.cell(row=1, column=(iteration+1)).value = iteration

    sheet.cell(row=1, column=1).value = "day"
    sheet.cell(row=2, column=1).value = "mf_active"
    sheet.cell(row=3, column=1).value = "mf_onsite"
    sheet.cell(row=4, column=1).value = "mf_remote"

    row_count = 4
    for each_location in range(0, (len(worksites))):
        for each_mf in range(0, worksites[each_location].mf_capacity):
            row_count += 1
            sheet.cell(row=row_count, column=1).value = worksites[each_location].name


def populate(data: dict) -> None:
    """
    Populates the sheet with data.
    :param data: dictionary with location assignment data
    :return: None
    """
    for each_entry in range(2, len(data)+1):

        # add titles for rows containing MFCount data
        sheet.cell(row=2, column=(each_entry+1)).value = data["day_" + str(each_entry)]["Onsite_Remote"].active
        sheet.cell(row=3, column=(each_entry+1)).value = data["day_" + str(each_entry)]["Onsite_Remote"].onsite
        sheet.cell(row=4, column=(each_entry+1)).value = data["day_" + str(each_entry)]["Onsite_Remote"].remote

        row_count = 4
        all_worksite_names = data["day_" + str(each_entry)].keys()

        for each_worksite in all_worksite_names:

            # hardcoded this part due to lack of time
            if each_worksite != "Onsite_Remote":
                if each_worksite == "MB_Bronx":
                    row_count = 4
                elif each_worksite == "MB_Pleasantville":
                    row_count = 5
                elif each_worksite == "LPA":
                    row_count = 9
                elif each_worksite == "BECA":
                    row_count = 11
                elif each_worksite == "ICHS":
                    row_count = 14
                elif each_worksite == "FIHS":
                    row_count = 16
                elif each_worksite == "Remote":
                    row_count = 19

                for each_person in data["day_" + str(each_entry)][each_worksite]:
                    row_count += 1
                    sheet.cell(row=row_count, column=(each_entry+1)).value = each_person

    # save file once writing is done
    workbook.save(filename=filename)
